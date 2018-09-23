# Версия программы 1.05 от 20.07.2018

import os
# сменить директорию
#os.chdir("D:\Задачи\ExcelConvertorTest")
# получить текущую рабочую директорию
cwd = os.getcwd()
# список файлов и директорий в текущей директории
filesInDir = os.listdir('.')

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from datetime import *
from PIL import Image as ImagePIL

print('Список объектов в исходной директории:')
print(filesInDir)
print()

while True:
    mainFileName = str(input('Введите полное название исходного файла: '))
    if os.path.exists(cwd + '/' + mainFileName):
        wbMain = load_workbook(mainFileName, data_only=True)
        wsMain = wbMain['all_rk']
        break
    else:
        print('Файл <<' + mainFileName + '>> не найден!')
        continue

# задаём шрифты, всего будут использоваться 2 шрифта
fontCalibri = Font(name='Calibri', size=8, bold=False, italic=False, vertAlign=None,
                   underline='none', strike=False, color='FF000000')

fontTNR = Font(name='Times New Roman', size=10, bold=False, italic=False,
               vertAlign=None, underline='none', strike=False, color='FF000000')

# устанавливаем формат границ (границы ячеек будут заданы при создании файла)
border = Border(left=Side(border_style='thin', color='FF000000'),
                   right=Side(border_style='thin', color='FF000000'),
                   top=Side(border_style='thin', color='FF000000'),
                   bottom=Side(border_style='thin', color='FF000000'),
                   diagonal=Side(border_style='thin', color='FF000000'),
                   diagonal_direction=0,
                   outline=Side(border_style='thin', color='FF000000'),
                   vertical=Side(border_style='thin', color='FF000000'),
                   horizontal=Side(border_style='thin', color='FF000000'))

# устанавливаем свойства текста ячеек
align_center = Alignment(horizontal='center', vertical='center', text_rotation=0,
                       wrap_text=True, shrink_to_fit=False, indent=0)

align_left = Alignment(horizontal='left', vertical='bottom', text_rotation=0,
                       wrap_text=False, shrink_to_fit=False, indent=0)

number_format = 'General'

# задаём ячейки с толстой правой границей
rightBoldBorderCells = ['A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11', 'A12', 
                       'A13', 'A14', 'A15', 'A16', 'A17', 'A18', 'A19', 'A20', 'A21', 'A22',   
                       'B2', 'B15', 'H6', 'H7', 'H8', 'H9', 'H10', 'H11', 'H12', 'H13',
                       'H14', 'H15', 'H20', 'H21', 'H22']

# задаём ячейки с толстой левой границей
leftBoldBorderCells = ['C3', 'C4', 'C5', 'C15', 'C16', 'C17', 'C18', 'C19', 'C20']

# задаём ячейки с толстой нижней границей
bottomBoldBorderCells = ['B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']

# задаём ячейки с толстой верхней границей
topBoldBorderCells = ['D2', 'E2', 'F2', 'G2', 'C6', 'D6', 'E6', 'F6', 'G6',
                      'C15', 'D15', 'E15', 'F15', 'G15', 'C20', 'D20', 'E20',
                      'F20', 'G20', 'B23', 'C23', 'D23', 'E23', 'F23', 'G23', 'H23']

# задаём ячейки с толстыми боковыми границами
sideBoldBorderCells = ['H2', 'H3', 'H4', 'H5', 'H15', 'H16', 'H17', 'H18', 'H19']

# задаём номера строк, в которых ячейки колонок (D:E) и (F:G) будут объединены
mergeCellsRowNumbers = ['2', '3', '4', '5']

# функция ищет колонку с определённым содержанием и возвращает её номер
def _find_cell_num_by_value(cellContainsValue):
    colNumber = 1
    try:
        while wsMain.cell(row = 1, column = colNumber).value != cellContainsValue:
            colNumber += 1
        return int(colNumber)
    except:
        print('Колонка <' + str(cellContainsValue) + '> в документе не найдена!')

# получаем номер колонкок с искомым заголовком
columnWithID = _find_cell_num_by_value('Сквозной идентификатор')
columnWithConstrType = _find_cell_num_by_value('Тип рекламной конструкции в альбоме схемы размещения')
columnWithNoScheme = _find_cell_num_by_value('Номер рекламной конструкции в альбоме схемы размещения')
columnWithAdress = _find_cell_num_by_value('Улица,магистраль')
columnWithOwner = _find_cell_num_by_value('Собственник')
columnWithPhotoNumber = _find_cell_num_by_value('number_photo')
columnWithPhoto1 = _find_cell_num_by_value('Photo1')
columnWithPhoto2 = _find_cell_num_by_value('Photo2')
columnWithPhoto3 = _find_cell_num_by_value('Photo3')
columnWithCoordX1_1 = _find_cell_num_by_value('X64_1')
columnWithCoordX1_2 = _find_cell_num_by_value('X64_2')
columnWithCoordX1_3 = _find_cell_num_by_value('X64_3')
columnWithCoordX1_4 = _find_cell_num_by_value('X64_4')
columnWithCoordX1_5 = _find_cell_num_by_value('X64_5')
columnWithCoordX1_6 = _find_cell_num_by_value('X64_6')
columnWithCoordY1_1 = _find_cell_num_by_value('Y64_1')
columnWithCoordY1_2 = _find_cell_num_by_value('Y64_2')
columnWithCoordY1_3 = _find_cell_num_by_value('Y64_3')
columnWithCoordY1_4 = _find_cell_num_by_value('Y64_4')
columnWithCoordY1_5 = _find_cell_num_by_value('Y64_5')
columnWithCoordY1_6 = _find_cell_num_by_value('Y64_6')

columnWithCoordX2_1 = _find_cell_num_by_value('X84_1')
columnWithCoordX2_2 = _find_cell_num_by_value('X84_2')
columnWithCoordX2_3 = _find_cell_num_by_value('X84_3')
columnWithCoordX2_4 = _find_cell_num_by_value('X84_4')
columnWithCoordX2_5 = _find_cell_num_by_value('X84_5')
columnWithCoordX2_6 = _find_cell_num_by_value('X84_6')
columnWithCoordY2_1 = _find_cell_num_by_value('Y84_1')
columnWithCoordY2_2 = _find_cell_num_by_value('Y84_2')
columnWithCoordY2_3 = _find_cell_num_by_value('Y84_3')
columnWithCoordY2_4 = _find_cell_num_by_value('Y84_4')
columnWithCoordY2_5 = _find_cell_num_by_value('Y84_5')
columnWithCoordY2_6 = _find_cell_num_by_value('Y84_6')

#создаём папки для хранения создаваемых файлов (временные картинки и конечный результат)
if not os.path.exists(cwd + '/temp_images/'):
    os.mkdir("temp_images")
if not os.path.exists(cwd + '/xlsx_result/'):
    os.mkdir("xlsx_result")

# определяем количество строк для обработки в исходном файле
row_count = wsMain.max_row

# задаём цикл по всем строкам исходного документа
# цикл генерирует новые файлы по шаблону и заполняет их данными из строк исходного файла
rowNumberForCycle = 2
while rowNumberForCycle <= row_count:
    
    if wsMain.cell(row=rowNumberForCycle, column=columnWithPhotoNumber).value > 0:

        wbNew = Workbook()
        wsNew = wbNew.active
        
        # рисуем толстые границы
        for rightBoldBorderCount in rightBoldBorderCells:
            wsNew[rightBoldBorderCount].border = Border(right=Side(border_style='medium', color='FF000000'))
        for leftBoldBorderCount in leftBoldBorderCells:
            wsNew[leftBoldBorderCount].border = Border(left=Side(border_style='medium', color='FF000000'))
        for topBoldBorderCount in topBoldBorderCells:
            wsNew[topBoldBorderCount].border = Border(top=Side(border_style='medium', color='FF000000'))
        for bottomBoldBorderCount in bottomBoldBorderCells:
            wsNew[bottomBoldBorderCount].border = Border(bottom=Side(border_style='medium', color='FF000000'))
        for sideBoldBorderCount in sideBoldBorderCells:
            wsNew[sideBoldBorderCount].border = Border(right=Side(border_style='medium', color='FF000000'),
                                                       left=Side(border_style='medium', color='FF000000'))

        #угловые жирные границы
        wsNew['C2'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                    left=Side(border_style='medium', color='FF000000'))
        wsNew['H2'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                    right=Side(border_style='medium', color='FF000000'),
                                    left=Side(border_style='medium', color='FF000000'))
         
        # устанавливаем высоту строк
        wsNew.row_dimensions[1].height = 2
        wsNew.row_dimensions[2].height = 64 
        wsNew.row_dimensions[3].height = 39
        wsNew.row_dimensions[4].height = 21
        wsNew.row_dimensions[5].height = 21
        wsNew.row_dimensions[6].height = 15
        wsNew.row_dimensions[7].height = 15
        wsNew.row_dimensions[8].height = 18
        wsNew.row_dimensions[9].height = 18
        wsNew.row_dimensions[10].height = 18
        wsNew.row_dimensions[11].height = 18
        wsNew.row_dimensions[12].height = 18
        wsNew.row_dimensions[13].height = 18


        # устанавливаем ширину колонок
        wsNew.column_dimensions['A'].width = 1
        wsNew.column_dimensions['B'].width = 2 
        wsNew.column_dimensions['C'].width = 30
        wsNew.column_dimensions['D'].width = 15
        wsNew.column_dimensions['E'].width = 15
        wsNew.column_dimensions['F'].width = 15
        wsNew.column_dimensions['G'].width = 15
        wsNew.column_dimensions['H'].width = 2

        # объединяем по 4 ячейки в колонке А
        wsNew.merge_cells('B2:B5') 
        wsNew['B2'].alignment = align_center
        wsNew['B2'].font = fontCalibri
        wsNew['B2'] = 'О п р е д е л е н и е'
    
        # попарно объединяем ячейки в колонках (D:E) и (F:G)
        for mrgPairsNo in mergeCellsRowNumbers:
            wsNew.merge_cells('D' + mrgPairsNo + ':E' + mrgPairsNo)
            wsNew.merge_cells('F' + mrgPairsNo + ':G' + mrgPairsNo)

        # выравнивание и шрифт для ячеек кроме первой колонки
        for cellObj in wsNew['C2:H23']:
            for cell in cellObj:
                wsNew[cell.coordinate].alignment = align_center
                wsNew[cell.coordinate].font = fontTNR

        # заполняем ячейки данными
        wsNew['C2'] = 'Единый идентификационный номер реклманой конструкции'
        wsNew['D2'] = 'Номер рекламной конструкции, согласно схеме размещения рекламных конструкций Муниципального района'
        wsNew['F2'] = 'Адресная характеристика рекламной конструкции, согласно схеме размещения рекламных конструкций Муниципального района'
        wsNew['C4'].alignment = align_left
        wsNew['C4'] = 'Собственник:'
        wsNew['C5'].alignment = align_left
        wsNew['C5'] = 'Тип рекламной конструкции:'
        wsNew['D7'] = 'X1'
        wsNew['E7'] = 'X2'
        wsNew['F7'] = 'Y1'
        wsNew['G7'] = 'Y2'

        # записываем в определённые ячейки нового файла данные
        # из соответствующих ячеек исходного файла
        wsNew['C3'] = wsMain.cell(row=rowNumberForCycle, column=columnWithID).value
        wsNew['D3'] = wsMain.cell(row=rowNumberForCycle, column=columnWithNoScheme).value
        wsNew['F3'] = wsMain.cell(row=rowNumberForCycle, column=columnWithAdress).value
        wsNew['D4'] = wsMain.cell(row=rowNumberForCycle, column=columnWithOwner).value
        wsNew['D5'] = wsMain.cell(row=rowNumberForCycle, column=columnWithConstrType).value
        wsNew['D8'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX1_1).value
        wsNew['F8'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY1_1).value
        wsNew['D9'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX1_2).value
        wsNew['F9'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY1_2).value
        wsNew['D10'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX1_3).value
        wsNew['F10'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY1_3).value
        wsNew['D11'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX1_4).value
        wsNew['F11'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY1_4).value
        wsNew['D12'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX1_5).value
        wsNew['F12'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY1_5).value
        wsNew['D13'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX1_6).value
        wsNew['F13'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY1_6).value

        wsNew['E8'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX2_1).value
        wsNew['G8'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY2_1).value
        wsNew['E9'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX2_2).value
        wsNew['G9'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY2_2).value
        wsNew['E10'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX2_3).value
        wsNew['G10'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY2_3).value
        wsNew['E11'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX2_4).value
        wsNew['G11'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY2_4).value
        wsNew['E12'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX2_5).value
        wsNew['G12'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY2_5).value
        wsNew['E13'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordX2_6).value
        wsNew['G13'] = wsMain.cell(row=rowNumberForCycle, column=columnWithCoordY2_6).value

        # название файла соответствует содержимому колонки ID в соответствующей строке
        newFileID = wsMain.cell(row=rowNumberForCycle, column=columnWithID).value

        # заменяем символы двоеточия, недопустимые в названии файла, на нижнее подчёркивание
        if newFileID is not None:
            changedDotsNewFileID = newFileID.replace(':', '_')
            changedNewFileID = changedDotsNewFileID.replace('/', '-')
        else:
            break
        wsNew.title = changedNewFileID
    
        fullBorderCells = ['C7', 'D7', 'E7', 'F7', 'G7']

        # удаляем пустые строки, в которых нет координат
        counter1 = 0
        counter2 = 0
        while counter1 < 6:
            if wsNew.cell(row=int(8 + counter1), column=4).value is not None and wsNew.cell(row=int(8 + counter1), column=4).value !='':
                wsNew['C'+str(8+counter1)] = str(1 + counter1)
                fullBorderCells.append('C' + str(8 + counter1))
                fullBorderCells.append('D' + str(8 + counter1))
                fullBorderCells.append('E' + str(8 + counter1))
                fullBorderCells.append('F' + str(8 + counter1))
                fullBorderCells.append('G' + str(8 + counter1))
            else:
                wsNew.delete_rows(int(8 + counter1), 1)
                counter2 += 1
            counter1 += 1

        wsNew['F'+str(21-counter2)].alignment = align_left
        wsNew['F'+str(21-counter2)] = 'Номер страницы'
        wsNew['G'+str(21-counter2)].alignment = align_center
        wsNew['G'+str(21-counter2)] = str(rowNumberForCycle - 1)
        wsNew['C'+str(16-counter2)] = 'Фото стороны A'
        wsNew['D'+str(18-counter2)] = 'Фото стороны C (при наличии)'
        wsNew['F'+str(16-counter2)] = 'Фото стороны B'
    
        wsNew.row_dimensions[14 - counter2].height = 12
        wsNew.row_dimensions[15 - counter2].height = 12
        wsNew.row_dimensions[16 - counter2].height = 15
        wsNew.row_dimensions[17 - counter2].height = 12
        wsNew.row_dimensions[18 - counter2].height = 18
        wsNew.row_dimensions[19 - counter2].height = 12
        wsNew.row_dimensions[20 - counter2].height = 12
        wsNew.row_dimensions[21 - counter2].height = 12
        wsNew.row_dimensions[22 - counter2].height = 12

        # задаём адрес фотографии на диске, копируя значения из требуемой ячейки
        imgAdresFull = 'P:/5_Дорнадзор/6_Проекты/00_Текущие проекты/636_РК_ЛО_2017/4. Обработанное/Геллер/150618/export img/'
        img1addres = wsMain.cell(row=rowNumberForCycle, column=columnWithPhoto1).value
        if img1addres is not None and len(str(img1addres))>0:
            try:
                img1 = ImagePIL.open(imgAdresFull + img1addres)
                # задаём базовое значение ширины фотографии
                basewidth1 = 210 
                wpercent = (basewidth1/float(img1.size[0]))
                hsize1 = int((float(img1.size[1])*float(wpercent)))
                img1small = img1.resize((basewidth1,hsize1), ImagePIL.LANCZOS)
                img1small.save(cwd + "/temp_images/" + changedNewFileID + '_p1.jpg')
            except:
                with open(cwd + "/xlsx_result/" + "_errorlog.txt", "a") as errorTXT:
                    errorTXT.write(str('Для файла ' + changedNewFileID + ' не удалось создать изображение <<Фотография 1>>.\n'))
                print('Для файла ' + changedNewFileID + ' не удалось создать изображение <<Фотография 1>>.')

        img2addres = wsMain.cell(row=rowNumberForCycle, column=columnWithPhoto2).value
        if img2addres is not None and len(str(img2addres))>0:
            try:
                img2 = ImagePIL.open(imgAdresFull+img2addres)
                basewidth2 = 210 
                wpercent = (basewidth2/float(img2.size[0]))
                hsize2 = int((float(img2.size[1])*float(wpercent)))
                img2small = img2.resize((basewidth2,hsize2), ImagePIL.LANCZOS)
                img2small.save(cwd + "/temp_images/" + changedNewFileID + '_p2.jpg')
            except:
                with open(cwd + "/xlsx_result/" + "_errorlog.txt", "a") as errorTXT:
                    errorTXT.write(str('Для файла ' + changedNewFileID + ' не удалось создать изображение <<Фотография 2>>.\n'))
                print('Для файла ' + changedNewFileID + ' не удалось создать изображение <<Фотография 2>>.')

        img3addres = wsMain.cell(row=rowNumberForCycle, column=columnWithPhoto3).value
        if img3addres is not None and len(str(img3addres))>0:
            try:
                img3 = ImagePIL.open(imgAdresFull+img3addres)
                basewidth3 = 210 
                wpercent = (basewidth3/float(img3.size[0]))
                hsize3 = int((float(img3.size[1])*float(wpercent)))
                img3small = img3.resize((basewidth3,hsize3), ImagePIL.LANCZOS)
                img3small.save(cwd + "/temp_images/" + changedNewFileID + '_p3.jpg')
            except:
                with open(cwd + "/xlsx_result/" + "_errorlog.txt", "a") as errorTXT:
                    errorTXT.write(str('Для файла ' + changedNewFileID + ' не удалось создать изображение <<Фотография 3>>.\n'))
                print('Для файла ' + changedNewFileID + ' не удалось создать изображение <<Фотография 3>>.')

        # вставляем изображения в соответствующие ячейки
        if os.path.exists(cwd + "/temp_images/" + changedNewFileID + '_p1.jpg'):
            img1 = Image(cwd + "/temp_images/" + changedNewFileID + '_p1.jpg')
            wsNew.add_image(img1, 'C'+str(16 - counter2))
            # устанавливаем высоту строки с фотографиями (она равна 0,75 от высоты фото в пикселях)
            row15height = max(hsize1, hsize2)
            wsNew.row_dimensions[16 - counter2].height = (row15height * 3)//4
        else:
            with open(cwd + "/xlsx_result/" + "_errorlog.txt", "a") as errorTXT:
                errorTXT.write(str('Для файла ' + changedNewFileID + ' изображение <<Фотография 1>> не обнаружено.\n'))
            print('Для файла ' + changedNewFileID + ' изображение <<Фотография 1>> не обнаружено.')

        if os.path.exists(cwd + "/temp_images/" + changedNewFileID + '_p2.jpg'):
            img2 = Image(cwd + "/temp_images/" + changedNewFileID + '_p2.jpg')
            wsNew.add_image(img2, 'F'+str(16 - counter2))
            if os.path.exists(cwd + "/temp_images/" + changedNewFileID + '_p1.jpg'):
                row15height = max(hsize1, hsize2)
            else:
                row15height = max(15, hsize2)
            wsNew.row_dimensions[16 - counter2].height = (row15height * 3)//4
        else:
            with open(cwd + "/xlsx_result/" + "_errorlog.txt", "a") as errorTXT:
                errorTXT.write(str('Для файла ' + changedNewFileID + ' изображение <<Фотография 2>> не обнаружено.\n'))
            print('Для файла ' + changedNewFileID + ' изображение <<Фотография 2>> не обнаружено.')

        if os.path.exists(cwd + "/temp_images/" + changedNewFileID + '_p3.jpg'):
            img3 = Image(cwd + "/temp_images/" + changedNewFileID + '_p3.jpg')
            wsNew.add_image(img3, 'D' + str(18 - counter2))
            row16height = hsize3
            wsNew.row_dimensions[18 - counter2].height = (row16height * 3)//4
        else:
            wsNew.row_dimensions[18 - counter2].height = 18
    
        # объединяем отдельные ячейки
        wsNew.merge_cells('D' + str(18 - counter2)+':E' + str(18 - counter2))
        wsNew.merge_cells('F' + str(16 - counter2)+':G' + str(16 - counter2))

        wsNew.merge_cells('B6:B'+str(14 - counter2))
        wsNew['B6'].alignment = align_center
        wsNew['B6'].font = fontCalibri
        wsNew['B6'] = 'К о о р д и н а т ы'

        wsNew.merge_cells('B' + str(15 - counter2) + ':B' + str(18 - counter2))
        wsNew['B'+ str(15 - counter2)].alignment = align_center
        wsNew['B'+ str(15 - counter2)].font = fontCalibri
        wsNew['B'+ str(15 - counter2)] = 'О п и с а н и е'

        # рисуем границы для ячеек, у которых есть все 4 границы
        for fullBorderCount in fullBorderCells:
            wsNew[fullBorderCount].border = Border(left=Side(border_style='thin', color='FF000000'),
                                                   right=Side(border_style='thin', color='FF000000'),
                                                   top=Side(border_style='thin', color='FF000000'),
                                                   bottom=Side(border_style='thin', color='FF000000'))
    
        wsNew['C'+str(15 - counter2)].border = Border(top=Side(border_style='medium', color='FF000000'), 
                                                      left=Side(border_style='medium', color='FF000000'))
        wsNew['G'+str(15 - counter2)].border = Border(top=Side(border_style='medium', color='FF000000'), 
                                                      right=Side(border_style='medium', color='FF000000'))
        wsNew['H'+str(16 - counter2)].border = Border(left=Side(border_style='medium', color='FF000000'), 
                                                      right=Side(border_style='medium', color='FF000000'))
        wsNew['C'+str(16 - counter2)].border = Border(left=Side(border_style='medium', color='FF000000'))
        wsNew['D'+str(15 - counter2)].border = Border(top=Side(border_style='medium', color='FF000000'))
        wsNew['E'+str(15 - counter2)].border = Border(top=Side(border_style='medium', color='FF000000'))
        wsNew['F'+str(15 - counter2)].border = Border(top=Side(border_style='medium', color='FF000000'))

        # сохраняем файл с названием, соответствующим Унифицированному ID
        if not os.path.exists(cwd + "/xlsx_result/" + changedNewFileID + ".xlsx"):
            wbNew.save(cwd + "/xlsx_result/" + changedNewFileID + ".xlsx")
            wbNew.close()
            print('ID ' + newFileID + ' обработан. Создан файл ' + str(rowNumberForCycle - 1) + ' из ' + str(row_count - 1))
        else:
            print('ID ' + newFileID + ' обработан. Файл ' + str(changedNewFileID) + ' уже существует!')

        # удаляем временные файлы изображений
        if os.path.exists(cwd + "/temp_images/" + changedNewFileID + '_p1.jpg'):
            os.remove(cwd + "/temp_images/" + changedNewFileID + '_p1.jpg')
        if os.path.exists(cwd + "/temp_images/" + changedNewFileID + '_p2.jpg'):
            os.remove(cwd + "/temp_images/" + changedNewFileID + '_p2.jpg')
        if os.path.exists(cwd + "/temp_images/" + changedNewFileID + '_p3.jpg'):
            os.remove(cwd + "/temp_images/" + changedNewFileID + '_p3.jpg')
    rowNumberForCycle += 1

# удаляем папку temp_images
os.rmdir(cwd + "/temp_images")

errorTXT.close()
wbMain.close()